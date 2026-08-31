"""Build IS110_gold_v0 from Durrant/Perry 2024 Nature supplementary materials.

Sources (unpacked under /global/scratch/users/kh36969/DL_novel_guide_editor/IS110_gold/unpacked/):
  Supp Table 1 : IS621 reference LE / RE / donor sequences + 274 ortholog phylogeny
  Supp Table 2 : 39 experimentally used bridge RNAs with explicit TBL/DBL specificity strings
  Supp Table 3 : per-insertion characterization (632 WT sites + 169 Programmed sites)
                 with paired (bRNA, actual genomic 11/14 bp target, DBL target, Levenshtein)
  Supp Table 5 : 274 IS110/IS1111 elements with group and 7 consensus bRNA structures
  Supp Data 1  : 5511 target–bRNA + 2201 donor–bRNA covariation MSA rows (224 cols)

This script emits ONE canonical JSONL of "pairing events":

  {
    system_id       : globally unique key (bRNA name + insertion index)
    ortholog_id     : bRNA name (from Table 2)
    group           : IS110 | IS1111 | unknown
    brna_sequence   : full RNA sequence of the bridge RNA (from Table 2)
    tbl_spec_11bp   : canonical TBL 11 bp specificity (Table 2)
    dbl_spec_11bp   : canonical DBL 11 bp specificity (Table 2)
    tbl_spec_14bp   : canonical TBL 14 bp specificity (Table 3, when present)
    dbl_spec_14bp   : canonical DBL 14 bp specificity (Table 3, when present)
    genome_target_11bp : observed genomic insertion site 11 bp (Table 3)
    genome_target_14bp : observed genomic insertion site 14 bp
    genome_core     : 2 bp junction core (CT / CG / ...)
    tbl_lev         : Levenshtein of TBL_spec vs genome_target (from Table 3)
    dbl_lev         : Levenshtein of DBL_spec vs canonical DBL target
    contig_id       : reference contig
    ins_start       : insertion start
    ins_end         : insertion end
    ins_strand      : + / -
    source          : 'Durrant2024_WT' | 'Durrant2024_Programmed'
    on_target       : bool (Programmed set annotates On-Target vs Off-Target)
    reads_bio_rep1  : experimental read count
    reads_bio_rep2  : experimental read count
    avg_reads_pct   : average insertion-reads % (proxy for pairing efficiency)
    rank            : within-bRNA insertion-site rank
  }

Downstream:
  scripts/audit_is110_gold_pairing.py runs the cognate-vs-shuffled benchmark
  on this JSONL, no HMM catalog / no Sniffles / no 350bp heuristic NC.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

BASE = Path('/global/scratch/users/kh36969/DL_novel_guide_editor/IS110_gold/unpacked')
OUT_DIR = Path('/global/scratch/users/kh36969/DL_novel_guide_editor/IS110_gold/curated')
OUT_DIR.mkdir(parents=True, exist_ok=True)

TABLE2 = BASE / '2023-09-16026B-s3' / '2023-09-16026B-SupplementaryTable2.xlsx'
TABLE3 = BASE / '2023-09-16026B-s3' / '2023-09-16026B-SupplementaryTable3.xlsx'
TABLE5 = BASE / '2023-09-16026B-s3' / '2023-09-16026B-SupplementaryTable5.xlsx'


def _find(header, sub):
    for i, c in enumerate(header):
        if c and sub.lower() in c.lower():
            return i
    return None


def load_table2():
    """Return {bRNA_name: {TBL, DBL, brna_seq}} for all 39 bRNAs."""
    wb = openpyxl.load_workbook(TABLE2, data_only=True)
    ws = wb['Key bridge RNAs Used']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    nm = _find(header, 'Name')
    tbl = _find(header, 'Target_Binding_Loop_Specificity')
    dbl = _find(header, 'Donor_Binding_Loop_Specificity')
    seq = _find(header, 'Sequence')
    out = {}
    for r in rows[1:]:
        if not r or r[nm] is None:
            continue
        name = r[nm]
        out[name] = {
            'TBL': None if not r[tbl] or r[tbl] == 'N/A' else r[tbl],
            'DBL': None if not r[dbl] or r[dbl] == 'N/A' else r[dbl],
            'brna_seq': r[seq] if r[seq] else None,
        }
    wb.close()
    return out


def match_brna(name, table2):
    """Match a Table 3 row's bRNA ID to Table 2's key (exact or fuzzy)."""
    if name in table2:
        return name
    for k in table2:
        if k in name or name in k:
            return k
    return None


def hamming(a, b):
    if not a or not b or len(a) != len(b):
        return None
    return sum(1 for x, y in zip(a, b) if x != y)


def load_table3(sheet_name, source_label, table2):
    wb = openpyxl.load_workbook(TABLE3, data_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    out = []

    def _col(sub): return _find(header, sub)
    bR_i = _col('Bridge RNA ID')
    contig_i = _col('Contig ID')
    st_i = _col('Insertion Core Start')
    en_i = _col('Insertion Core End')
    strand_i = _col('Insertion Strand')
    core_i = _col('Genome Core Sequence')
    gs11_i = _col('Genome Insertion Site Sequence 11bp')
    gs14_i = _col('Genome Insertion Site Sequence 14bp')
    tt11_i = _col('TBL Target 11bp')
    tt14_i = _col('TBL Target 14bp')
    dt11_i = _col('DBL Target 11bp')
    dt14_i = _col('DBL Target 14bp')
    tlev_i = _col('TBL Specificity Levenshtein Distance')
    dlev_i = _col('DBL Specificity Levenshtein Distance')
    rep1_i = _col('Bio. Rep. 1 Read Count')
    rep2_i = _col('Bio. Rep. 2 Read Count')
    pct_i = _col('Avg. Insertion Reads (%)')
    rank_i = _col('Insertion Site Rank')
    ontgt_i = _col('Off-target Category')
    rtg_i = _col('RTG Status')

    def _v(r, i):
        return r[i] if i is not None else None

    for r in rows[1:]:
        if r is None or r[bR_i] is None:
            continue
        bR = r[bR_i]
        match = match_brna(bR, table2)
        specs = table2.get(match, {}) if match else {}
        rec = {
            'source': source_label,
            'bR_id_raw': bR,
            'ortholog_id': match,
            'brna_sequence': specs.get('brna_seq'),
            'tbl_spec_11bp': specs.get('TBL'),
            'dbl_spec_11bp': specs.get('DBL'),
            'tbl_spec_14bp': _v(r, tt14_i),
            'dbl_spec_14bp': _v(r, dt14_i),
            'genome_target_11bp': _v(r, gs11_i),
            'genome_target_14bp': _v(r, gs14_i),
            'genome_core': _v(r, core_i),
            'tbl_lev': _v(r, tlev_i),
            'dbl_lev': _v(r, dlev_i),
            'contig_id': _v(r, contig_i),
            'ins_start': _v(r, st_i),
            'ins_end': _v(r, en_i),
            'ins_strand': _v(r, strand_i),
            'on_target': (_v(r, ontgt_i) == 'On-Target') if ontgt_i is not None else None,
            'rtg_status': _v(r, rtg_i),
            'reads_bio_rep1': _v(r, rep1_i),
            'reads_bio_rep2': _v(r, rep2_i),
            'avg_reads_pct': _v(r, pct_i),
            'rank': _v(r, rank_i),
        }
        # Recompute Levenshtein when missing but seqs present
        if rec['tbl_lev'] is None and rec['tbl_spec_11bp'] and rec['genome_target_11bp']:
            rec['tbl_lev'] = hamming(rec['tbl_spec_11bp'], rec['genome_target_11bp'])
        out.append(rec)
    wb.close()
    return out


def load_table5():
    """Load 274 IS110/IS1111 element phylogeny (element → group)."""
    wb = openpyxl.load_workbook(TABLE5, data_only=True)
    ws = wb['Fig. 6C Phylogenetic Tree Info']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    nm_i = _find(header, 'IS Element Name')
    grp_i = _find(header, 'IS110 Group')
    end_i = _find(header, 'Bridge RNA End')
    out = {}
    for r in rows[1:]:
        if r is None or r[nm_i] is None:
            continue
        out[r[nm_i]] = {'group': r[grp_i], 'bR_end': r[end_i]}
    wb.close()
    return out


def main():
    print(f'[load] Supp Table 2 → bridge RNA specs')
    t2 = load_table2()
    print(f'  {len(t2)} bRNAs; {sum(1 for v in t2.values() if v["TBL"])} with TBL spec, '
          f'{sum(1 for v in t2.values() if v["DBL"])} with DBL spec')

    print(f'[load] Supp Table 5 → 274-element phylogeny')
    t5 = load_table5()
    from collections import Counter
    print(f'  {len(t5)} elements; group dist = {Counter(v["group"] for v in t5.values())}')

    print(f'[load] Supp Table 3 → per-insertion records')
    wt = load_table3('WT Genomic Insertion Sites', 'Durrant2024_WT', t2)
    pg = load_table3('Programmed Genomic Ins. Sites', 'Durrant2024_Programmed', t2)
    print(f'  WT sheet: {len(wt)} records')
    print(f'  Programmed sheet: {len(pg)} records')

    # Attach system_id + group. For each record, group = t5.get(bR name → element name)
    # (bRNA IDs like "bridge_RNA_T-WT_D-WT" don't map directly to IS elements in Table 5,
    #  since all 39 bRNAs are derived from IS621. Assign group="IS110" for the bridge RNAs
    #  used in this paper -- Durrant confirms IS621 is IS110 group.)
    all_rows = []
    for i, rec in enumerate(wt + pg):
        rec['system_id'] = f'{rec["source"]}_{i:05d}'
        rec['group'] = 'IS110'  # per Table 5, IS621 = IS110 group
        all_rows.append(rec)
    print(f'  total emitted: {len(all_rows)} rows')

    out_path = OUT_DIR / 'is110_gold_v0.jsonl'
    with out_path.open('w') as f:
        for r in all_rows:
            f.write(json.dumps(r) + '\n')
    print(f'\n[write] {out_path}   ({out_path.stat().st_size} bytes)')

    # Quick numerical summary
    n_with_tbl = sum(1 for r in all_rows if r['tbl_spec_11bp'] and r['genome_target_11bp'])
    n_with_dbl = sum(1 for r in all_rows if r['dbl_spec_11bp'] and r['genome_target_11bp'])
    n_wt_ontgt = sum(1 for r in all_rows if r['source'] == 'Durrant2024_WT')
    n_pg_ontgt = sum(1 for r in all_rows if r['source'] == 'Durrant2024_Programmed' and r['on_target'])
    print(f'\n  usable TBL-arm rows (both specs present): {n_with_tbl}')
    print(f'  usable DBL-arm rows: {n_with_dbl}')
    print(f'  WT (all cognate insertions): {n_wt_ontgt}')
    print(f'  Programmed on-target: {n_pg_ontgt}')


if __name__ == '__main__':
    main()
